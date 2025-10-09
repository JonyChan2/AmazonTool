#import pyautogui as gui
#import keyboard as kbd
import pyperclip
from time import sleep
import json
import openpyxl as pxl
from tkinter import filedialog
from openpyxl.formatting.rule import FormatObject

filename = filedialog.askopenfilename(
    title='选择json文件',
    defaultextension=".json",
    filetypes=[('JSON File(.json)','.json'), ("All files", ".*")]
    )
print(filename.split('/')[-1].split('.')[0])
js = open(filename,'r',encoding='utf-8').readlines()
js_line = ''
for line in js:
    js_line = js_line + line
js_serial = json.loads(js_line)

excelfile = filedialog.askopenfilename(
    title='选择Excel文件',
    defaultextension=".xlsx",
    filetypes=[('Excel File(.xlsx)','.xlsx'), ("All files", ".*")]
)
wb = pxl.load_workbook(excelfile)
ws1 = wb.create_sheet(filename.split('/')[-1].split('.')[0])
js_serial = js_serial['data']['items']
row = len(js_serial)

ws1.cell(1,1,'关键词')
ws1.cell(1,2,'流量占比')
ws1.cell(1,3,'广告占比')
ws1.cell(1,4,'排名位置/页数')
ws1.cell(1,5,'ABA占比')
ws1.cell(1,6,'SPR')
ws1.cell(1,7,'PPC竞价')

#新建一个字典，将变量名和json里的变量名关联起来

for i in range(row):
    try:
        keyword = js_serial[i]["keywords"]
        ws1.cell(i+2,1,keyword)
        traffic_pct = js_serial[i]["trafficPercentage"]
        ws1.cell(i+2,2,traffic_pct).number_format = '0.00%'
        ad_pct = js_serial[i]["adRatio"]
        ws1.cell(i+2,3,ad_pct).number_format = '0.00%'
        rank_pos = js_serial[i]["rankPosition"]["position"]
        rank_pg = js_serial[i]["rankPosition"]["page"]
        ws1.cell(i+2,4,'{0}/{1}'.format(rank_pos,rank_pg))
        aba = js_serial[i]["monopolyClickRate"]
        ws1.cell(i+2,5,aba).number_format = '0.00%'
        spr = js_serial[i]["titleDensityExact"]
        ws1.cell(i+2,6,spr)
        bid = js_serial[i]["bid"]
        ws1.cell(i+2,7,bid).number_format = '$0.00'
    except:
        continue
#print(js_serial['list'])

wb.save(excelfile)