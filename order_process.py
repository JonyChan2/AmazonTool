import openpyxl as pxl
import datetime
import tkinter as tk
from tkinter import filedialog
import pyautogui as gui
from time import sleep

purchaseDate = []
orderNumber = []

def order_download():
    logocord = gui.locateOnScreen('img/ziniaologo.png')
    x, y = gui.center(logocord)
    gui.click(x,y)
    gui.hotkey('ctrl','t')
    gui.press('tab',3)
    gui.write('https://sellercentral.amazon.com/reportcentral/FlatFileAllOrdersReport/1')
    gui.press('enter')
    sleep(2)
    gui.hotkey('ctrl','0')
    while (1):
        try:
            gui.click(gui.center(gui.locateOnScreen('img/lastday.png')))
            break
        except:
            print('not found, retrying...')
            sleep(1)
    gui.press('down')
    gui.press('enter')
    gui.click(gui.center(gui.locateOnScreen('img/ask_download.png')))
    gui.keyDown('shift')
    gui.scroll(-400)
    gui.keyUp('shift')
    gui.scroll(-400)
    print('Wait for it to finish',end='')
    sleep(1)
    wfd = 'Waiting for download'
    while (1):
        try:
            down = gui.locateOnScreen('img/file_download.png')
            break
        except:
            sleep(1)
            wfd += '.'
            print('\r'+wfd,end='')

    x, y = gui.center(down)
    gui.click(x-10,y+15)
    sleep(2)
    gui.press('enter')
    print('Download Complete!')
#od.self()
def format():
    filepath = filedialog.askopenfilename()
    file = open(filepath,'r',encoding='utf-8')
    lines = file.readlines()[1:-1]
    orders = []
    for line in lines:
        cells = []
        for cell in line.split('\t'):
            if cell != '':
                cells.append(cell)
        orders.append(cells)
    orders.reverse()    
    return orders

def order_process():
    data_stream = format()
    wb = pxl.Workbook()
    ws = wb['Sheet']
    row = 1
    for column in data_stream:
        if column[11] == 'Cancelled':
            continue
        date = column[2][0:10].split('-')
        ws.cell(row,1,datetime.datetime(int(date[0]),int(date[1]),int(date[2]))).number_format = "yyyy/m/d" #date
        ws.cell(row,3,column[1]) #order-number
        ws.cell(row,5,column[10]) #asin
        ws.cell(row,6,int(column[12])) #quantity
        ws.cell(row,7,column[11]) #order-state

        row+=1
    wb.save(filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[('Excel Files(.xlsx)','.xlsx'), ("All files", ".*")]
        )
    )

data_stream = []
window = tk.Tk()
window.geometry('300x200')
window.title('AmazonOrderProcessor')

title = tk.Label(text="It's ugly, but it's useful")
export_butt = tk.Button(text='1.export orders to txt',command=order_download)
process_butt = tk.Button(text='2.txt to excel',command=order_process)
merge_title = tk.Label(text='3.Merge new record to old one manually.')
butt = tk.Button(text='4.')

title.pack()
export_butt.pack()
process_butt.pack()
merge_title.pack()
butt.pack()
window.mainloop()