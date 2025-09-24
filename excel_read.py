import openpyxl as pxl
import tkinter as tk
from tkinter import filedialog
import pyautogui as gui
import tripstone as ts
import pyperclip

def wb_search():
    print(f'Looking for:{order.get()}')
    cord_result = []
    # Iterate through all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Iterate through rows and columns
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if str(cell.value) == order.get(): # Replace with your search term
                    print(f"Found '{cell.value}' in sheet '{sheet}', coordinate: {cell.coordinate}")
                    cord_result.append([cell.column_letter,cell.row])
    rows = []
    for cell in ws[cord_result[0][1]]:
        rows.append(cell.value)
    result.set(rows)

def contact():
    customer_name = ''
    order = ''
    ts.new_tab()
    ts.paste('https://sellercentral.amazon.com/orders-v3/order/' + pyperclip.paste())
    gui.press('enter')
    return

def open_xlsx():
    file_path = filedialog.askopenfilename(
        title="Select a File",
        initialdir="/",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    return file_path

def ws_index():
    ws_index = []
    for row in wb['Sheet'].iter_rows(values_only=0):
        ws_index.append(row)
    return ws_index

wb = pxl.load_workbook('sample.xlsx')

window = tk.Tk()
window.title('Order Processor')
window.geometry('400x200')

order = tk.StringVar()
result = tk.StringVar()
order.set('')
result.set('')

text = tk.Label(window,text='Last order number:',anchor='w').pack()
last_order = tk.Entry(window,textvariable=order).pack()
submit_btn = tk.Button(window,text='Submit',command=wb_search).pack()
start_btn = tk.Button(window,text='Start Customer Contact').pack()
search_result = tk.Label(window,textvariable=result).pack()
window.mainloop()